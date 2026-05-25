from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, Response
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
import sqlite3
import qrcode
from io import BytesIO
import os
import shutil
import uuid
from contextlib import contextmanager

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="Restaurant POS System")

os.makedirs("receipts", exist_ok=True)
os.makedirs("static", exist_ok=True)
os.makedirs("exports", exist_ok=True)
os.makedirs("static/uploads", exist_ok=True)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATABASE = "pos_system.db"

@contextmanager
def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()

# ─── Excel helper ──────────────────────────────────────────────────────────

EXCEL_FILE = "exports/orders_export.xlsx"
EXCEL_HEADERS = [
    'Order #', 'Table', 'Date & Time', 'Items', 'Subtotal (RS)',
    'Tax (5%)', 'Total (RS)', 'Payment Method', 'Payment Status', 'Note'
]
HEADER_COLOR = "1a1a2e"
PAID_BG = "d1fae5"; PAID_FG = "065f46"
PENDING_BG = "fed7aa"; PENDING_FG = "92400e"
DELETED_BG = "fee2e2"; DELETED_FG = "991b1b"

def _thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def _get_or_create_workbook():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders Report"
        hfont = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        for col, h in enumerate(EXCEL_HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = 'A2'
    order_rows = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            order_rows[str(val)] = row
    return wb, ws, order_rows

def _write_order_row(ws, row_idx, order_number, table_number, created_at,
                     items_display, subtotal, tax, total,
                     payment_method, payment_status, note=""):
    border = _thin_border()
    row_data = [
        order_number, f"Table {table_number}", created_at, items_display,
        float(subtotal), float(tax), float(total),
        payment_method or 'N/A', payment_status.upper(), note
    ]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        if col_idx == 9:
            status_upper = str(value).upper()
            if status_upper == 'PAID':
                cell.fill = PatternFill(start_color=PAID_BG, end_color=PAID_BG, fill_type="solid")
                cell.font = Font(color=PAID_FG, bold=True)
            elif status_upper == 'DELETED':
                cell.fill = PatternFill(start_color=DELETED_BG, end_color=DELETED_BG, fill_type="solid")
                cell.font = Font(color=DELETED_FG, bold=True)
            else:
                cell.fill = PatternFill(start_color=PENDING_BG, end_color=PENDING_BG, fill_type="solid")
                cell.font = Font(color=PENDING_FG, bold=True)
        if note == "DELETED":
            cell.font = Font(strike=True, color=DELETED_FG, bold=(col_idx == 9))

def _auto_width(ws):
    for column in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

def _items_display_for_order(db, order_id):
    items = db.execute("""
        SELECT m.name, oi.quantity FROM order_items oi
        JOIN menu_items m ON oi.menu_item_id = m.id WHERE oi.order_id = ?
    """, (order_id,)).fetchall()
    return ", ".join(f"{i['name']} x{i['quantity']}" for i in items) if items else "—"

def excel_upsert_order(order_id: int, note: str = ""):
    try:
        with get_db() as db:
            order = db.execute("""
                SELECT o.*, t.table_number FROM orders o
                JOIN tables t ON o.table_id = t.id WHERE o.id = ?
            """, (order_id,)).fetchone()
            if not order:
                return
            items_display = _items_display_for_order(db, order_id)
        wb, ws, order_rows = _get_or_create_workbook()
        order_num_str = str(order['order_number'])
        row_idx = order_rows.get(order_num_str, ws.max_row + 1)
        _write_order_row(ws, row_idx, order['order_number'], order['table_number'],
                         order['created_at'], items_display, order['subtotal'], order['tax'],
                         order['total'], order['payment_method'], order['payment_status'], note)
        _auto_width(ws)
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"⚠️ Excel upsert failed: {e}")

def excel_mark_deleted(order_id: int):
    try:
        with get_db() as db:
            order = db.execute("""
                SELECT o.*, t.table_number FROM orders o
                JOIN tables t ON o.table_id = t.id WHERE o.id = ?
            """, (order_id,)).fetchone()
            if not order:
                return
            items_display = _items_display_for_order(db, order_id)
        wb, ws, order_rows = _get_or_create_workbook()
        order_num_str = str(order['order_number'])
        row_idx = order_rows.get(order_num_str, ws.max_row + 1)
        _write_order_row(ws, row_idx, order['order_number'], order['table_number'],
                         order['created_at'], items_display, order['subtotal'], order['tax'],
                         order['total'], order['payment_method'], "DELETED", "DELETED")
        _auto_width(ws)
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"⚠️ Excel mark-deleted failed: {e}")

# ─── Database init ─────────────────────────────────────────────────────────

def init_database():
    with get_db() as db:
        db.execute("""CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL, icon TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS menu_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, price DECIMAL(10,2) NOT NULL,
            category_id INTEGER, description TEXT, icon TEXT,
            image_url TEXT, is_available BOOLEAN DEFAULT 1,
            stock_quantity INTEGER DEFAULT 999,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (category_id) REFERENCES categories(id))""")
        try:
            db.execute("ALTER TABLE menu_items ADD COLUMN image_url TEXT")
        except:
            pass
        db.execute("""CREATE TABLE IF NOT EXISTS tables (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_number INTEGER UNIQUE NOT NULL,
            capacity INTEGER DEFAULT 4, status TEXT DEFAULT 'available')""")
        db.execute("""CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_id INTEGER, order_number TEXT UNIQUE,
            status TEXT DEFAULT 'pending',
            subtotal DECIMAL(10,2), tax DECIMAL(10,2), total DECIMAL(10,2),
            payment_method TEXT, payment_status TEXT DEFAULT 'unpaid',
            customer_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            FOREIGN KEY (table_id) REFERENCES tables(id))""")
        db.execute("""CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER, menu_item_id INTEGER,
            quantity INTEGER, unit_price DECIMAL(10,2),
            subtotal DECIMAL(10,2), notes TEXT,
            FOREIGN KEY (order_id) REFERENCES orders(id),
            FOREIGN KEY (menu_item_id) REFERENCES menu_items(id))""")

        cursor = db.execute("SELECT COUNT(*) as count FROM categories")
        if cursor.fetchone()['count'] == 0:
            cats = [('Burgers','🍔'),('Pizzas','🍕'),('Salads','🥗'),
                    ('Appetizers','🍤'),('Beverages','🥤'),('Desserts','🍰'),('Mains','🍽️')]
            for c in cats:
                db.execute("INSERT INTO categories (name, icon) VALUES (?, ?)", c)
            items = [
                ('Classic Cheeseburger',12.99,1,'Angus beef, cheddar cheese','🍔',999),
                ('Spicy Chicken Burger',11.49,1,'Crispy chicken with spicy sauce','🍗',999),
                ('Margherita Pizza',14.99,2,'Fresh mozzarella, basil','🍕',999),
                ('Pepperoni Feast',16.49,2,'Double pepperoni, extra cheese','🍕',999),
                ('Greek Salad',8.99,3,'Feta, olives, cucumber','🥗',999),
                ('Crispy Calamari',9.49,4,'With lemon aioli','🦑',999),
                ('Loaded Nachos',8.99,4,'Cheese, jalapeños','🌮',999),
                ('Craft Soda',3.49,5,'Cola or Ginger Ale','🥤',999),
                ('Iced Latte',4.99,5,'With oat milk option','☕',999),
                ('Molten Lava Cake',6.99,6,'With vanilla ice cream','🍰',999),
                ('Grilled Salmon',21.99,7,'Lemon butter sauce','🐟',999),
                ('Steak Frites',24.99,7,'Sirloin steak with fries','🥩',999),
            ]
            for i in items:
                db.execute("""INSERT INTO menu_items (name,price,category_id,description,icon,stock_quantity)
                    VALUES (?,?,?,?,?,?)""", i)
            for n in range(1, 11):
                db.execute("INSERT INTO tables (table_number, capacity) VALUES (?, ?)", (n, 4))

init_database()

# ─── Pydantic Models ────────────────────────────────────────────────────────

class OrderItemCreate(BaseModel):
    menu_item_id: int
    quantity: int
    notes: Optional[str] = ""

class OrderCreate(BaseModel):
    table_id: int
    items: List[OrderItemCreate]
    customer_name: Optional[str] = ""

class PaymentRequest(BaseModel):
    payment_method: str
    amount: float

class TableCreate(BaseModel):
    table_number: int
    capacity: Optional[int] = 4

class MenuItemCreate(BaseModel):
    name: str
    price: float
    category_id: int
    description: Optional[str] = ""
    icon: Optional[str] = "🍽️"
    image_url: Optional[str] = ""

class MenuItemUpdate(BaseModel):
    name: str
    price: float
    category_id: int
    description: Optional[str] = ""
    icon: Optional[str] = "🍽️"
    image_url: Optional[str] = ""

class AvailabilityUpdate(BaseModel):
    is_available: int

class CategoryCreate(BaseModel):
    name: str
    icon: Optional[str] = "📁"

# ─── Image Upload ──────────────────────────────────────────────────────────

@app.post("/api/upload-image")
async def upload_image(file: UploadFile = File(...)):
    try:
        allowed = {"image/jpeg", "image/png", "image/gif", "image/webp"}
        if file.content_type not in allowed:
            return {"success": False, "message": "Only JPG/PNG/GIF/WEBP allowed"}
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "jpg"
        filename = f"{uuid.uuid4().hex}.{ext}"
        dest = os.path.join("static", "uploads", filename)
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        return {"success": True, "url": f"/static/uploads/{filename}"}
    except Exception as e:
        return {"success": False, "message": str(e)}

# ─── Categories ────────────────────────────────────────────────────────────

@app.get("/api/categories")
def get_categories():
    with get_db() as db:
        rows = db.execute("SELECT * FROM categories ORDER BY name").fetchall()
        return [dict(r) for r in rows]

@app.post("/api/admin/categories")
async def admin_add_category(request: Request):
    try:
        data = await request.json()
        name = data.get('name', '').strip()
        icon = data.get('icon', '📁')
        if not name:
            return {"success": False, "message": "Category name is required"}
        with get_db() as db:
            if db.execute("SELECT id FROM categories WHERE name = ?", (name,)).fetchone():
                return {"success": False, "message": f"Category '{name}' already exists"}
            db.execute("INSERT INTO categories (name, icon) VALUES (?, ?)", (name, icon))
        return {"success": True, "message": f"Category '{name}' added successfully"}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.delete("/api/admin/categories/{category_id}")
def admin_delete_category(category_id: int):
    try:
        with get_db() as db:
            cat = db.execute("SELECT id, name FROM categories WHERE id = ?", (category_id,)).fetchone()
            if not cat:
                return {"success": False, "message": "Category not found"}
            db.execute("UPDATE menu_items SET category_id = NULL WHERE category_id = ?", (category_id,))
            db.execute("DELETE FROM categories WHERE id = ?", (category_id,))
        return {"success": True, "message": f"Category '{cat['name']}' deleted"}
    except Exception as e:
        return {"success": False, "message": str(e)}

# ─── Menu Items ────────────────────────────────────────────────────────────

@app.get("/api/menu-items")
def get_menu_items(category_id: Optional[int] = None):
    with get_db() as db:
        query = """SELECT m.*, c.name as category_name
            FROM menu_items m LEFT JOIN categories c ON m.category_id = c.id
            WHERE m.is_available = 1 AND m.stock_quantity > 0"""
        params = []
        if category_id:
            query += " AND m.category_id = ?"
            params.append(category_id)
        query += " ORDER BY m.name"
        return [dict(r) for r in db.execute(query, params).fetchall()]

@app.get("/api/admin/menu-items")
def admin_get_all_menu_items():
    with get_db() as db:
        rows = db.execute("""SELECT m.*, c.name as category_name
            FROM menu_items m LEFT JOIN categories c ON m.category_id = c.id
            ORDER BY m.name""").fetchall()
        result = []
        for r in rows:
            d = dict(r); d['price'] = float(d['price'])
            result.append(d)
        return result

@app.get("/api/admin/menu-items/{item_id}")
def admin_get_menu_item(item_id: int):
    with get_db() as db:
        item = db.execute("""SELECT m.*, c.name as category_name
            FROM menu_items m LEFT JOIN categories c ON m.category_id = c.id
            WHERE m.id = ?""", (item_id,)).fetchone()
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        r = dict(item); r['price'] = float(r['price'])
        return r

@app.post("/api/admin/menu-items")
def admin_add_menu_item(item: MenuItemCreate):
    try:
        with get_db() as db:
            if not db.execute("SELECT id FROM categories WHERE id = ?", (item.category_id,)).fetchone():
                return {"success": False, "message": "Category not found"}
            db.execute("""INSERT INTO menu_items (name,price,category_id,description,icon,image_url,is_available,stock_quantity)
                VALUES (?,?,?,?,?,?,1,999)""",
                (item.name, item.price, item.category_id, item.description, item.icon, item.image_url or ""))
        return {"success": True, "message": f'"{item.name}" added to menu'}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.put("/api/admin/menu-items/{item_id}")
def admin_update_menu_item(item_id: int, item: MenuItemUpdate):
    try:
        with get_db() as db:
            if not db.execute("SELECT id FROM menu_items WHERE id = ?", (item_id,)).fetchone():
                return {"success": False, "message": "Item not found"}
            db.execute("""UPDATE menu_items SET name=?,price=?,category_id=?,description=?,icon=?,image_url=?
                WHERE id=?""", (item.name, item.price, item.category_id, item.description, item.icon, item.image_url or "", item_id))
        return {"success": True, "message": f'"{item.name}" updated'}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.delete("/api/admin/menu-items/{item_id}")
def admin_delete_menu_item(item_id: int):
    try:
        with get_db() as db:
            item = db.execute("SELECT id, name FROM menu_items WHERE id = ?", (item_id,)).fetchone()
            if not item:
                return {"success": False, "message": "Item not found"}
            db.execute("UPDATE menu_items SET is_available=0, stock_quantity=0 WHERE id=?", (item_id,))
        return {"success": True, "message": f'"{item["name"]}" removed from menu'}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.put("/api/admin/menu-items/{item_id}/availability")
def admin_toggle_availability(item_id: int, data: AvailabilityUpdate):
    try:
        with get_db() as db:
            if not db.execute("SELECT id FROM menu_items WHERE id = ?", (item_id,)).fetchone():
                return {"success": False, "message": "Item not found"}
            if data.is_available:
                db.execute("UPDATE menu_items SET is_available=1, stock_quantity=999 WHERE id=?", (item_id,))
            else:
                db.execute("UPDATE menu_items SET is_available=0 WHERE id=?", (item_id,))
        return {"success": True, "message": "Availability updated"}
    except Exception as e:
        return {"success": False, "message": str(e)}

# ─── Tables ────────────────────────────────────────────────────────────────

@app.get("/api/tables")
def get_tables():
    with get_db() as db:
        rows = db.execute("""SELECT t.*,
            CASE WHEN o.id IS NOT NULL AND o.payment_status='unpaid' THEN 'occupied'
                 ELSE 'available' END as current_status
            FROM tables t
            LEFT JOIN orders o ON t.id=o.table_id AND o.payment_status='unpaid'
            ORDER BY t.table_number""").fetchall()
        return [dict(r) for r in rows]

@app.post("/api/admin/tables")
def admin_add_table(table: TableCreate):
    try:
        with get_db() as db:
            if db.execute("SELECT id FROM tables WHERE table_number=?", (table.table_number,)).fetchone():
                return {"success": False, "message": f"Table {table.table_number} already exists"}
            db.execute("INSERT INTO tables (table_number, capacity) VALUES (?, ?)", (table.table_number, table.capacity))
        return {"success": True, "message": f"Table {table.table_number} added"}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.delete("/api/admin/tables/{table_id}")
def admin_delete_table(table_id: int):
    try:
        with get_db() as db:
            table = db.execute("SELECT id, table_number FROM tables WHERE id=?", (table_id,)).fetchone()
            if not table:
                return {"success": False, "message": "Table not found"}
            if db.execute("SELECT id FROM orders WHERE table_id=? AND payment_status='unpaid'", (table_id,)).fetchone():
                return {"success": False, "message": "Cannot delete table with active unpaid orders"}
            db.execute("DELETE FROM tables WHERE id=?", (table_id,))
        return {"success": True, "message": f"Table {table['table_number']} deleted"}
    except Exception as e:
        return {"success": False, "message": str(e)}

# ─── Orders ────────────────────────────────────────────────────────────────

@app.post("/api/orders")
def create_order(order: OrderCreate):
    with get_db() as db:
        order_number = f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        subtotal = 0
        items_data = []
        for item in order.items:
            mi = db.execute("SELECT price FROM menu_items WHERE id=?", (item.menu_item_id,)).fetchone()
            if not mi:
                raise HTTPException(status_code=404, detail="Menu item not found")
            s = mi['price'] * item.quantity
            subtotal += s
            items_data.append({'menu_item_id': item.menu_item_id, 'quantity': item.quantity,
                                'unit_price': mi['price'], 'subtotal': s, 'notes': item.notes})
        tax = round(subtotal * 0.05, 2)
        total = round(subtotal + tax, 2)
        cur = db.execute("""INSERT INTO orders (table_id,order_number,subtotal,tax,total,customer_name)
            VALUES (?,?,?,?,?,?)""",
            (order.table_id, order_number, subtotal, tax, total, order.customer_name))
        order_id = cur.lastrowid
        for d in items_data:
            db.execute("""INSERT INTO order_items (order_id,menu_item_id,quantity,unit_price,subtotal,notes)
                VALUES (?,?,?,?,?,?)""",
                (order_id, d['menu_item_id'], d['quantity'], d['unit_price'], d['subtotal'], d['notes']))
            db.execute("UPDATE menu_items SET stock_quantity=stock_quantity-? WHERE id=?",
                       (d['quantity'], d['menu_item_id']))
    excel_upsert_order(order_id, note="")
    return {"order_id": order_id, "order_number": order_number, "total": total}

@app.get("/api/orders")
def get_orders(status: Optional[str] = None):
    with get_db() as db:
        query = """SELECT o.*, t.table_number,
            COUNT(oi.id) as item_count,
            COALESCE(SUM(oi.quantity),0) as total_items
            FROM orders o JOIN tables t ON o.table_id=t.id
            LEFT JOIN order_items oi ON o.id=oi.order_id WHERE 1=1"""
        params = []
        if status:
            query += " AND o.status=?"; params.append(status)
        query += " GROUP BY o.id ORDER BY o.created_at DESC"
        result = []
        for o in db.execute(query, params).fetchall():
            d = dict(o)
            d['total'] = float(d['total'])
            d['subtotal'] = float(d['subtotal'])
            d['tax'] = float(d['tax'])
            result.append(d)
        return result

@app.get("/api/orders/{order_id}")
def get_order(order_id: int):
    with get_db() as db:
        order = db.execute("""SELECT o.*, t.table_number FROM orders o
            JOIN tables t ON o.table_id=t.id WHERE o.id=?""", (order_id,)).fetchone()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        items = db.execute("""SELECT oi.*, m.name, m.icon FROM order_items oi
            JOIN menu_items m ON oi.menu_item_id=m.id WHERE oi.order_id=?""", (order_id,)).fetchall()
        result = dict(order)
        result['total'] = float(result['total'])
        result['subtotal'] = float(result['subtotal'])
        result['tax'] = float(result['tax'])
        result['items'] = []
        for i in items:
            d = dict(i)
            d['subtotal'] = float(d['subtotal'])
            d['unit_price'] = float(d['unit_price'])
            result['items'].append(d)
        return result

@app.delete("/api/orders/{order_id}")
def delete_order(order_id: int):
    try:
        excel_mark_deleted(order_id)
        with get_db() as db:
            if not db.execute("SELECT id FROM orders WHERE id=?", (order_id,)).fetchone():
                return {"success": False, "message": "Order not found"}
            db.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
            db.execute("DELETE FROM orders WHERE id=?", (order_id,))
        return {"success": True, "message": f"Order #{order_id} deleted"}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/orders/delete-multiple")
async def delete_multiple_orders(request: Request):
    try:
        data = await request.json()
        order_ids = data.get('order_ids', [])
        if not order_ids:
            return {"success": False, "message": "No order IDs provided"}
        for oid in order_ids:
            excel_mark_deleted(oid)
        with get_db() as db:
            for oid in order_ids:
                db.execute("DELETE FROM order_items WHERE order_id=?", (oid,))
                db.execute("DELETE FROM orders WHERE id=?", (oid,))
        return {"success": True, "message": f"Deleted {len(order_ids)} order(s)"}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/orders/clear-all")
async def clear_all_orders():
    try:
        with get_db() as db:
            order_ids = [r['id'] for r in db.execute("SELECT id FROM orders").fetchall()]
        for oid in order_ids:
            excel_mark_deleted(oid)
        with get_db() as db:
            db.execute("DELETE FROM order_items")
            db.execute("DELETE FROM orders")
            try:
                db.execute("DELETE FROM sqlite_sequence WHERE name='orders'")
                db.execute("DELETE FROM sqlite_sequence WHERE name='order_items'")
            except:
                pass
        return {"success": True, "message": "All orders cleared"}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/orders/{order_id}/payment")
def process_payment(order_id: int, payment: PaymentRequest):
    with get_db() as db:
        order = db.execute("SELECT total, payment_status, table_id FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        if order['payment_status'] == 'paid':
            raise HTTPException(status_code=400, detail="Order already paid")
        db.execute("""UPDATE orders SET payment_method=?, payment_status='paid',
            status='completed', completed_at=CURRENT_TIMESTAMP WHERE id=?""",
            (payment.payment_method, order_id))
        db.execute("UPDATE tables SET status='available' WHERE id=?", (order['table_id'],))
    excel_upsert_order(order_id, note="")
    return {"message": "Payment processed successfully"}

# ─── Receipt ──────────────────────────────────────────────────────────────

@app.get("/api/orders/{order_id}/receipt")
def get_receipt(order_id: int):
    with get_db() as db:
        order = db.execute("""SELECT o.*, t.table_number FROM orders o
            JOIN tables t ON o.table_id=t.id WHERE o.id=?""", (order_id,)).fetchone()
        if not order:
            return HTMLResponse(content="<h1>Order not found</h1>", status_code=404)
        items = db.execute("""SELECT oi.*, m.name FROM order_items oi
            JOIN menu_items m ON oi.menu_item_id=m.id WHERE oi.order_id=?""", (order_id,)).fetchall()
    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Receipt #{order['order_number']}</title>
    <style>
        *{{margin:0;padding:0;box-sizing:border-box;}}
        body{{font-family:'Courier New',monospace;background:#fff;color:#000;}}
        .receipt{{width:320px;margin:0 auto;padding:24px 20px;}}
        .logo{{text-align:center;margin-bottom:16px;}}
        .logo h1{{font-size:22px;font-weight:900;letter-spacing:3px;}}
        .logo p{{font-size:11px;color:#666;margin-top:2px;}}
        .divider{{border:none;border-top:1px dashed #999;margin:12px 0;}}
        .info-row{{display:flex;justify-content:space-between;font-size:12px;margin:3px 0;}}
        .items{{margin:12px 0;}}
        .item{{display:flex;justify-content:space-between;font-size:13px;margin:6px 0;}}
        .item-name{{flex:1;}}
        .item-qty{{color:#666;margin:0 8px;}}
        .item-price{{font-weight:600;}}
        .totals{{margin-top:12px;}}
        .total-row{{display:flex;justify-content:space-between;font-size:13px;margin:4px 0;}}
        .grand-total{{font-size:16px;font-weight:900;border-top:2px solid #000;padding-top:8px;margin-top:8px;}}
        .footer{{text-align:center;margin-top:20px;font-size:11px;color:#666;line-height:1.6;}}
        @media print{{body{{margin:0;}}}}
    </style></head><body>
    <div class="receipt">
        <div class="logo">
            <h1>FUSION</h1>
            <p>Premium Restaurant Experience</p>
            <p>123 Food Street, Kathmandu</p>
            <p>Tel: +977-1-4XXXXXX</p>
        </div>
        <hr class="divider">
        <div class="info-row"><span>Order:</span><span><b>{order['order_number']}</b></span></div>
        <div class="info-row"><span>Table:</span><span>Table {order['table_number']}</span></div>
        <div class="info-row"><span>Date:</span><span>{datetime.now().strftime('%d/%m/%Y %H:%M')}</span></div>
        {"<div class='info-row'><span>Customer:</span><span>" + order['customer_name'] + "</span></div>" if order['customer_name'] else ""}
        <hr class="divider">
        <div class="items">"""
    for i in items:
        html += f"""<div class="item">
            <span class="item-name">{i['name']}</span>
            <span class="item-qty">x{i['quantity']}</span>
            <span class="item-price">RS {float(i['subtotal']):.2f}</span>
        </div>"""
    html += f"""</div>
        <hr class="divider">
        <div class="totals">
            <div class="total-row"><span>Subtotal</span><span>RS {float(order['subtotal']):.2f}</span></div>
            <div class="total-row"><span>Tax (5%)</span><span>RS {float(order['tax']):.2f}</span></div>
            <div class="total-row grand-total"><span>TOTAL</span><span>RS {float(order['total']):.2f}</span></div>
        </div>
        <hr class="divider">
        <div class="footer">
            <p>✦ Thank you for dining with us! ✦</p>
            <p>Please visit again</p>
            <p style="margin-top:8px;font-size:10px;">Powered by Fusion POS</p>
        </div>
    </div>
    <script>window.onload=function(){{window.print();}};</script>
    </body></html>"""
    return HTMLResponse(content=html)

@app.get("/api/payment/qrcode/{order_id}")
def generate_qr_code(order_id: int):
    with get_db() as db:
        order = db.execute("SELECT order_number, total FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
    esewa_url = f"https://esewa.com.np/epay/main?tAmt={float(order['total']):.2f}&amt={float(order['total']):.2f}&txAmt=0&psc=0&pdc=0&scd=RESTAURANTPOS&pid={order['order_number']}&su=http://localhost:8000&fu=http://localhost:8000"
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(esewa_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#60d394", back_color="white")
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return Response(content=buf.getvalue(), media_type="image/png")

@app.get("/api/export/excel")
def download_excel():
    if not os.path.exists(EXCEL_FILE):
        raise HTTPException(status_code=404, detail="No export file found. Create some orders first.")
    return FileResponse(EXCEL_FILE, filename="orders_export.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

app.mount("/static", StaticFiles(directory="static"), name="static_files")
app.mount("/", StaticFiles(directory="static", html=True), name="static")